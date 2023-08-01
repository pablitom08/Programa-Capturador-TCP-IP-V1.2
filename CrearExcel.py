#!/usr/bin/python
# -*- coding: utf-8 -*-
import openpyxl
# Creamos un nuevo archivo Excel
#import datetime
from datetime import datetime
import glob, os
import time
import getpass
import win32api

import win32con, win32gui, os

def check_if_file_is_open(filepath):
    if os.path.exists(filepath):
        try:
            os.rename(filepath, filepath)
            return False
        except OSError:
            return True
    else:
        return False


#today1 = datetime.now().date()

def registro(ide, directorio):
    apoyodir = getpass.getuser()
    today1 = datetime.now().date()
    path_today = f"C:/Users/{apoyodir}/Documents/Informes/{today1}.xlsx"

    #print("punto de control 2")

    if directorio == "":
        path_today = str(today1) + ".xlsx"
    else:
        path_today = directorio + "/" + str(today1) + ".xlsx"

 # Verificamos si el archivo está abierto
    while check_if_file_is_open(path_today):
        time.sleep(1)  # Esperamos un segundo antes de volver a verificar


    #print("punto de control 3")
    try:
        workbook = openpyxl.load_workbook(filename=path_today)

        # Obtenemos la hoja activa
        sheet = workbook.active

        # Obtenemos el número de la última fila no vacía
        last_row = sheet.max_row
        row = last_row + 1

    except:
        # Creamos un archivo Excel nuevo
        print(f"Creando el archivo de excel{path_today}")
        workbook = openpyxl.Workbook()

        # Obtenemos la hoja activa
        sheet = workbook.active

        # Escribimos algunos datos de ejemplo
        sheet['A1'] = 'FECHA ACTUAL'
        sheet['B1'] = 'LOTE'
        sheet['C1'] = 'NOMBRE DE PRODUCTO'
        sheet['D1'] = 'PESO'
        sheet['E1'] = 'UNIDAD'
        sheet['F1'] = 'RESPONSABLE'


        # Comenzamos a escribir desde la segunda fila
        row = 2

    # Escribimos los datos en la hoja de cálculo
    values = ide.split('*')
    sheet.cell(row=row, column=1, value=values[0])
    sheet.cell(row=row, column=2, value=values[1])
    sheet.cell(row=row, column=3, value=values[2])
    sheet.cell(row=row, column=4, value=values[3])
    sheet.cell(row=row, column=5, value=values[4])
    sheet.cell(row=row, column=6, value=values[5])

    # Incrementamos el número de fila
    row += 1

    # Guardamos el archivo Excel
    workbook.save(path_today)
    workbook.close()


if __name__ == '__main__':
    apoyodir=getpass.getuser()
    actual ="C:/Users/"+apoyodir+"/Documents"
    datos1=('aaa*eeee*234*brasil*123e*NH-34')
    registro(datos1,actual)
